import pandas as pd
from openpyxl import load_workbook

# Step 1: Define the file path for the Excel workbook
file_path = '../data/ReportsDB.xlsx'

# Step 2: Load the Excel workbook and the 'FinalReport' sheet
workbook = load_workbook(file_path)
ws = workbook['summary_report']

# Step 3: Read the necessary sheets into DataFrames
db_ul_extended_df = pd.read_excel(file_path, sheet_name='DB_U-L_extended')
add_extended_df = pd.read_excel(file_path, sheet_name='Add_extended')

# Step 4: Set the Site Code as index for faster lookups
add_extended_df.set_index('Site Code', inplace=True)

# Step 6: Set the column headers for the summary_report sheet
ws['A1'] = 'Start Date'
ws['B1'] = 'Start Time'
ws['C1'] = 'End Date'
ws['D1'] = 'End Time'
ws['E1'] = 'Duration'
ws['F1'] = 'Site Code'
ws['G1'] = 'Site_Code_Address_rus'
ws['H1'] = 'Site_Code_Address_eng'
ws['I1'] = 'ProblemDescription_eng'
ws['J1'] = 'ProblemDescription_rus'
ws['K1'] = 'ID'

# Step 7: Define start row in `summary_report` to begin inserting data
start_row = 2  # Start filling in data from row 5 (after headers)

# Step 8: Loop through the rows of `db_ul_extended_df` and insert the data into `summary_report`
for row_index in range(len(db_ul_extended_df)):
    row = db_ul_extended_df.iloc[row_index]  # Get the row based on index

    # Calculate the Excel row number directly
    row_num = start_row + row_index

    # Retrieve corresponding data from `Add_extended` using the site code as index
    site_code = row['Site Code']

    site_code_address_rus = add_extended_df.at[site_code, 'Site_Code_Address_rus']
    site_code_address_eng = add_extended_df.at[site_code, 'Site_Code_Address_eng']

    # Insert data into the corresponding cells
    ws[f'A{row_num}'] = row['Start Data for Luso']  # Start Date
    ws[f'B{row_num}'] = row['Start Time']  # Start Time
    ws[f'C{row_num}'] = row['End Date for Luso']  # End Date
    ws[f'D{row_num}'] = row['End Time']  # End Time
    ws[f'E{row_num}'] = row['Duration']  # Duration
    ws[f'F{row_num}'] = site_code  # Site Code
    ws[f'G{row_num}'] = site_code_address_rus  # Site_Code_Address_rus
    ws[f'H{row_num}'] = site_code_address_eng  # Site_Code_Address_eng
    ws[f'I{row_num}'] = row['ProblemDescription_eng']  # ProblemDescription_eng
    ws[f'J{row_num}'] = row['ProblemDescription_rus']  # ProblemDescription_rus

    # Get the ID from DB_U-L_extended
    id_value = row['ID']  # Assuming the ID is in the DB_U-L_extended sheet as well

    ws[f'K{row_num}'] = id_value  # ID

# Step 9: Save the workbook with the populated data
workbook.save(file_path)

print("Data has been successfully added to the 'summary_report' sheet.")
