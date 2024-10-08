import os
import pandas as pd
from openpyxl import Workbook


def get_filtered_data(start_date, end_date):
    """
    Function to filter data from the ReportsDB.xlsx file's sheet 'copy_summary_report'.
    """
    # Load the data from ReportsDB.xlsx
    source_filename = '../data/ReportsDB.xlsx'
    report_data = pd.read_excel(source_filename, sheet_name='copy_summary_report', header=0)

    # Convert the date columns to datetime
    report_data['Start Date'] = pd.to_datetime(report_data['Start Date'], errors='coerce')
    report_data['End Date'] = pd.to_datetime(report_data['End Date'], errors='coerce')

    # Filter data based on the provided dates
    filtered_data = report_data[
        (report_data['Start Date'] >= start_date) &
        (report_data['End Date'] <= end_date)
    ]

    # Exclude the ID column
    filtered_data = filtered_data.drop(columns=['ID'], errors='ignore')

    return filtered_data


def get_week_number(start_date, end_date, weeks_file='../data/ReportsDB.xlsx'):
    """
    Function to get the week number from the Weeks sheet of the ReportsDB.xlsx file.
    """
    try:
        # Load the Weeks sheet
        weeks_df = pd.read_excel(weeks_file, sheet_name='Weeks')

        # Convert the From_Date and To_Date columns to datetime format
        weeks_df['From_Date'] = pd.to_datetime(weeks_df['From_Date'], format='%d.%m.%Y')
        weeks_df['To_Date'] = pd.to_datetime(weeks_df['To_Date'], format='%d.%m.%Y')

        # Check if the input dates match exactly with any of the From_Date and To_Date in the sheet
        match = weeks_df[(weeks_df['From_Date'] == start_date) & (weeks_df['To_Date'] == end_date)]

        # Return the corresponding Week_number if a match is found
        if not match.empty:
            return match['Week_numbers'].values[0]
        else:
            return "Week XX"
    except FileNotFoundError:
        print(f"Error: The file '{weeks_file}' does not exist.")
        return "Week XX"
    except Exception as e:
        print(f"An error occurred while processing the Weeks data: {e}")
        return "Week XX"


def generate_report(start_date, end_date, filtered_data):
    # Format the dates for display
    formatted_start_date = start_date.strftime('%d-%m-%Y')
    formatted_end_date = end_date.strftime('%d-%m-%Y')

    # Load Weeks data from ReportsDB.xlsx
    source_filename = '../data/ReportsDB.xlsx'
    weeks_data = pd.read_excel(source_filename, sheet_name='Weeks', header=0)

    # Convert the From_Date and To_Date columns to datetime
    weeks_data['From_Date'] = pd.to_datetime(weeks_data['From_Date'], format='%d-%m-%Y', errors='coerce')
    weeks_data['To_Date'] = pd.to_datetime(weeks_data['To_Date'], format='%d-%m-%Y', errors='coerce')

    # Try to find a matching week based on the entered dates
    matching_week = weeks_data[
        (weeks_data['From_Date'] == start_date) &
        (weeks_data['To_Date'] == end_date)
    ]

    # Determine the week number
    week_number = "Week XX"  # Default if no match
    if not matching_week.empty:
        week_number = matching_week['Week_numbers'].values[0]

    # Create a new workbook for the report
    wb = Workbook()

    # Updated file name format
    report_filename = (f"Outage_Report_{start_date.year}_{start_date.month:02}_"
                       f"{start_date.day:02}_to_{end_date.year}_{end_date.month:02}_{end_date.day:02}.xlsx")

    # Define the directory to save the report
    save_directory = "../Weekly_Outage_Reports"
    os.makedirs(save_directory, exist_ok=True)

    # Save the report in the specified directory
    report_filepath = os.path.join(save_directory, report_filename)

    # Create a new sheet
    ws_merged = wb.active
    ws_merged.title = "Outages"

    # Add additional header rows before the main data
    additional_headers = [
        ["Weekly Report", "Technical outages and network failures"],
        ["Country", "XXX", "", "Period",
         f"{formatted_start_date} - {formatted_end_date}", week_number]  # Use the week number here
    ]

    # Append these rows to the new merged sheet
    for header_row in additional_headers:
        ws_merged.append(header_row)

    # Insert a blank row after the custom header
    ws_merged.append([])

    # Define the headers for the actual table data
    headers = [
        "Start Date", "Start Time", "End Date", "End Time", "Duration",
        "Site Code", "Site_Code_Address_rus", "Site_Code_Address_eng",
        "ProblemDescription_eng", "ProblemDescription_rus"
    ]

    # Append the headers to the merged sheet
    ws_merged.append(headers)

    # Prepare data for merging and format date columns
    formatted_data = []
    for row in filtered_data.values.tolist():
        # Convert Start Date and End Date to string
        start_date_str = row[0].strftime('%d-%m-%Y') if isinstance(row[0], pd.Timestamp) else row[0]
        end_date_str = row[2].strftime('%d-%m-%Y') if isinstance(row[2], pd.Timestamp) else row[2]

        # Convert Start Time and End Time to string
        start_time_str = row[1].strftime('%I:%M:%S %p') if isinstance(row[1], pd.Timestamp) else row[1]
        end_time_str = row[3].strftime('%I:%M:%S %p') if isinstance(row[3], pd.Timestamp) else row[3]

        formatted_row = [
            start_date_str,  # Formatted Start Date
            start_time_str,  # Formatted Start Time
            end_date_str,    # Formatted End Date
            end_time_str,    # Formatted End Time
            row[4],          # Duration (keep as is)
            row[5],          # Site Code
            row[6],          # Site_Code_Address_rus
            row[7],          # Site_Code_Address_eng
            row[8],          # ProblemDescription_eng
            row[9],          # ProblemDescription_rus
        ]
        formatted_data.append(formatted_row)

    # Create a dictionary to hold rows grouped by "ProblemDescription_eng"
    grouping_dict = {}
    for row in formatted_data:
        key = row[8]  # Use "ProblemDescription_eng" as the grouping key
        if key not in grouping_dict:
            grouping_dict[key] = []
        grouping_dict[key].append(row)

    # Start tracking rows for merged sheet
    current_row = len(additional_headers) + 3

    # Write grouped data into the new sheet and apply merging where needed
    for key, rows in grouping_dict.items():
        # Write the first row of the group
        ws_merged.append(rows[0])

        # Determine the starting row for this group
        start_row = current_row

        # Write additional rows in the same group
        for i in range(1, len(rows)):
            ws_merged.append(rows[i])

        # Merge cells for ProblemDescription_eng (column I, index 8)
        ws_merged.merge_cells(start_row=start_row, start_column=9, end_row=current_row + len(rows) - 1, end_column=9)

        # Merge cells for ProblemDescription_rus (column J, index 9)
        ws_merged.merge_cells(start_row=start_row, start_column=10, end_row=current_row + len(rows) - 1, end_column=10)

        # Update the current row for the next set of data
        current_row += len(rows)

    # Save the report to the specified directory
    wb.save(report_filepath)

    print(f"Created {report_filepath}")


def main():
    start_date_input = input("Start_Date (dd-mm-yyyy): ")
    end_date_input = input("End_Date (dd-mm-yyyy): ")

    # Convert string input to datetime using the correct format
    start_date = pd.to_datetime(start_date_input, format='%d-%m-%Y')
    end_date = pd.to_datetime(end_date_input, format='%d-%m-%Y')

    filtered_data = get_filtered_data(start_date, end_date)

    if not filtered_data.empty:
        generate_report(start_date, end_date, filtered_data)
    else:
        print("No data found for the specified date range.")


if __name__ == "__main__":
    main()
